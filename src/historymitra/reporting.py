from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUMMARY_HEADER_FONT = Font(color="FFFFFF", bold=True)
TABLE_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TABLE_HEADER_FONT = Font(bold=True)


def bool_label(value: Any) -> str:
    if value in (True, "1", 1, "true", "True", "YA", "Ya", "ya"):
        return "Ya"
    if value in (False, "0", 0, "false", "False", "", None):
        return "Tidak"
    return str(value)


def gender_label(value: Any) -> str:
    mapping = {
        "1": "Laki-laki",
        "2": "Perempuan",
    }
    return mapping.get(str(value), safe_text(value))


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def normalize_mitra_row(raw: dict[str, Any]) -> dict[str, str]:
    detail = raw.get("mitra_detail") or {}
    return {
        "id_mitra": safe_text(raw.get("id_mitra") or detail.get("id_mitra")),
        "sobat_id": safe_text(raw.get("sobat_id") or detail.get("sobat_id")),
        "nama_lengkap": safe_text(detail.get("nama_lengkap")),
        "nik_masked": safe_text(raw.get("nik") or detail.get("nik")),
        "email": safe_text(detail.get("email")),
        "no_telp": safe_text(detail.get("no_telp")),
        "nama_satker": safe_text(raw.get("nama_satker")),
        "nama_pos_rekrutmen": safe_text(raw.get("nama_pos")),
        "nama_pos_daftar_rekrutmen": safe_text(raw.get("nama_pos_daftar")),
        "ket_status_rekrutmen": safe_text(raw.get("ket_status")),
        "kd_survei_rekrutmen": safe_text(raw.get("kd_survei")),
        "id_kegiatan_rekrutmen": safe_text(raw.get("id_kegiatan")),
        "is_nik_verified": bool_label(raw.get("is_nik_verified")),
        "nik_verified_at": safe_text(raw.get("nik_verified_at")),
        "agama": safe_text(detail.get("agama")),
        "jenis_kelamin": gender_label(detail.get("jns_kelamin")),
        "tgl_lahir": safe_text(detail.get("tgl_lahir")),
        "pendidikan": safe_text(detail.get("pendidikan")),
        "pekerjaan": safe_text(detail.get("pekerjaan")),
        "status_kawin": safe_text(detail.get("status_kawin")),
        "alamat_prov": safe_text(detail.get("alamat_prov")),
        "alamat_kab": safe_text(detail.get("alamat_kab")),
        "alamat_kec": safe_text(detail.get("alamat_kec")),
        "alamat_desa": safe_text(detail.get("alamat_desa")),
        "alamat_detail": safe_text(detail.get("alamat_detail")),
        "is_bisa_komputer": bool_label(detail.get("is_bisa_komputer")),
        "is_capi": bool_label(detail.get("is_capi")),
        "is_hp_android": bool_label(detail.get("is_hp_android")),
        "is_laptop": bool_label(detail.get("is_laptop")),
        "is_motor": bool_label(detail.get("is_motor")),
        "is_naik_motor": bool_label(detail.get("is_naik_motor")),
        "catatan": safe_text(detail.get("catatan")),
        "created_at": safe_text(raw.get("CreatedAt")),
        "updated_at": safe_text(raw.get("UpdatedAt")),
    }


def enrich_mitra_row_with_detail(mitra_row: dict[str, str], detail: dict[str, Any]) -> dict[str, str]:
    enriched = dict(mitra_row)
    enriched.update(
        {
            "status_akun": safe_text(detail.get("status")),
            "alamat_is": bool_label(detail.get("alamat_is")),
            "is_pendataan_bps": bool_label(detail.get("is_pendataan_bps")),
            "is_sp": bool_label(detail.get("is_sp")),
            "is_st": bool_label(detail.get("is_st")),
            "is_se": bool_label(detail.get("is_se")),
            "is_susenas": bool_label(detail.get("is_susenas")),
            "is_sakernas": bool_label(detail.get("is_sakernas")),
            "is_sbh": bool_label(detail.get("is_sbh")),
            "merk_hp": safe_text(detail.get("merk_hp")),
            "tipe_hp": safe_text(detail.get("tipe_hp")),
            "ram_hp_gb": safe_text(detail.get("ram_hp")),
            "rekening_nama": safe_text(detail.get("rekening_nama")),
            "kd_bank": safe_text(detail.get("kd_bank")),
        }
    )
    return enriched


def normalize_history_documents(
    *,
    id_mitra: str,
    mitra_name: str,
    sobat_id: str,
    history_documents: list[dict[str, Any]],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for doc_index, doc in enumerate(history_documents, start=1):
        smds = doc.get("smds", [])
        source_label = "tahun_jalan" if doc_index == 1 else "histori_sebelumnya"
        for order, item in enumerate(smds, start=1):
            rows.append(
                {
                    "id_mitra": safe_text(id_mitra),
                    "nama_lengkap": safe_text(mitra_name),
                    "sobat_id": safe_text(sobat_id),
                    "history_source": source_label,
                    "history_doc_index": str(doc_index),
                    "history_order": str(order),
                    "id_ms": safe_text(item.get("id_ms")),
                    "kd_survei": safe_text(item.get("kd_survei")),
                    "nama_survei": safe_text(item.get("nama_survei")),
                    "id_keg": safe_text(item.get("id_keg")),
                    "nama_keg": safe_text(item.get("nama_keg")),
                    "id_pos": safe_text(item.get("id_pos")),
                    "nama_pos": safe_text(item.get("nama_pos")),
                    "id_pos_daftar": safe_text(item.get("id_pos_daftar")),
                    "nama_pos_daftar": safe_text(item.get("nama_pos_daftar")),
                    "kd_prov": safe_text(item.get("kd_prov")),
                    "nama_prov": safe_text(item.get("nama_prov")),
                    "kd_kab": safe_text(item.get("kd_kab")),
                    "nama_kab": safe_text(item.get("nama_kab")),
                    "wilayah_histori": " / ".join(
                        part for part in [safe_text(item.get("nama_prov")), safe_text(item.get("nama_kab"))] if part
                    ),
                    "status_kode_history": safe_text(item.get("status")),
                    "nama_status": safe_text(item.get("nama_status")),
                    "status_keg_kode": safe_text(item.get("status_keg")),
                    "status_survei_kode": safe_text(item.get("status_survei")),
                    "is_cocard_published": bool_label(item.get("is_cocard_published")),
                    "kd_mitra": safe_text(item.get("kd_mitra")),
                }
            )
    return rows


def build_rekap_mitra(history_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    bucket: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in history_rows:
        bucket[row["id_mitra"]].append(row)

    rekap_rows: list[dict[str, str]] = []
    for id_mitra, rows in bucket.items():
        surveys = [row["kd_survei"] for row in rows if row["kd_survei"]]
        unique_surveys = sorted(set(surveys))
        names = [row["nama_lengkap"] for row in rows if row["nama_lengkap"]]
        sobat_ids = [row["sobat_id"] for row in rows if row["sobat_id"]]
        latest_row = rows[0]
        rekap_rows.append(
            {
                "id_mitra": id_mitra,
                "nama_lengkap": names[0] if names else "",
                "sobat_id": sobat_ids[0] if sobat_ids else "",
                "total_histori": str(len(rows)),
                "total_survei_unik": str(len(unique_surveys)),
                "daftar_kd_survei": ", ".join(unique_surveys),
                "survei_terakhir_terlihat": latest_row["nama_survei"],
                "kegiatan_terakhir_terlihat": latest_row["nama_keg"],
                "status_terakhir_terlihat": latest_row["nama_status"],
            }
        )

    rekap_rows.sort(key=lambda item: (-int(item["total_histori"]), item["nama_lengkap"], item["id_mitra"]))
    return rekap_rows


def build_rekap_survey(history_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    survey_bucket: dict[str, dict[str, Any]] = {}
    for row in history_rows:
        key = row["kd_survei"]
        if key not in survey_bucket:
            survey_bucket[key] = {
                "kd_survei": row["kd_survei"],
                "nama_survei": row["nama_survei"],
                "total_histori": 0,
                "mitra_ids": set(),
                "status_counts": Counter(),
            }
        survey_bucket[key]["total_histori"] += 1
        survey_bucket[key]["mitra_ids"].add(row["id_mitra"])
        survey_bucket[key]["status_counts"][row["nama_status"]] += 1

    rows: list[dict[str, str]] = []
    for value in survey_bucket.values():
        status_summary = ", ".join(
            f"{status}: {count}" for status, count in value["status_counts"].most_common()
        )
        rows.append(
            {
                "kd_survei": value["kd_survei"],
                "nama_survei": value["nama_survei"],
                "total_histori": str(value["total_histori"]),
                "total_mitra_unik": str(len(value["mitra_ids"])),
                "ringkasan_status": status_summary,
            }
        )
    rows.sort(key=lambda item: (-int(item["total_histori"]), item["kd_survei"]))
    return rows


@dataclass(slots=True)
class ReportBundle:
    mitra_rows: list[dict[str, str]]
    history_rows: list[dict[str, str]]
    rekap_mitra_rows: list[dict[str, str]]
    rekap_survey_rows: list[dict[str, str]]
    source_label: str
    output_dir: Path


def format_terminal_summary(bundle: ReportBundle) -> str:
    total_mitra = len(bundle.mitra_rows)
    total_history = len(bundle.history_rows)
    mitra_with_history = len({row["id_mitra"] for row in bundle.history_rows})
    avg_history = mean([int(row["total_histori"]) for row in bundle.rekap_mitra_rows]) if bundle.rekap_mitra_rows else 0
    survey_counter = Counter(row["kd_survei"] for row in bundle.history_rows if row["kd_survei"])
    lines = [
        "",
        "=== Ringkasan History Mitra ===",
        f"Sumber data          : {bundle.source_label}",
        f"Folder output        : {bundle.output_dir}",
        f"Total mitra          : {total_mitra}",
        f"Mitra dengan histori : {mitra_with_history}",
        f"Total baris histori  : {total_history}",
        f"Rata-rata histori    : {avg_history:.2f} per mitra",
        f"Total survei unik    : {len(survey_counter)}",
        "",
        "Top survei:",
    ]
    if survey_counter:
        for index, (survey_code, count) in enumerate(survey_counter.most_common(10), start=1):
            lines.append(f"{index:>2}. {survey_code:<15} {count:>4} histori")
    else:
        lines.append("- Tidak ada data histori")
    return "\n".join(lines)


def autosize_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def write_table_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet(title)
    if not rows:
        worksheet["A1"] = "Tidak ada data"
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = TABLE_HEADER_FILL
        cell.font = TABLE_HEADER_FONT
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_worksheet_columns(worksheet)


def write_summary_sheet(
    workbook: Workbook,
    *,
    source_label: str,
    mitra_rows: list[dict[str, str]],
    history_rows: list[dict[str, str]],
    rekap_mitra_rows: list[dict[str, str]],
    rekap_survey_rows: list[dict[str, str]],
) -> None:
    worksheet = workbook.active
    worksheet.title = "Ringkasan"

    total_mitra = len(mitra_rows)
    total_history = len(history_rows)
    mitra_with_history = len({row["id_mitra"] for row in history_rows})
    avg_history = mean([int(row["total_histori"]) for row in rekap_mitra_rows]) if rekap_mitra_rows else 0

    summary_rows = [
        ("Dibuat pada", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Sumber data", source_label),
        ("Total mitra", total_mitra),
        ("Mitra dengan histori", mitra_with_history),
        ("Total baris histori", total_history),
        ("Total survei unik", len(rekap_survey_rows)),
        ("Rata-rata histori per mitra", round(avg_history, 2)),
    ]

    worksheet["A1"] = "Ringkasan History Mitra BPS Bulungan"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.merge_cells("A1:B1")

    row_index = 3
    for label, value in summary_rows:
        worksheet.cell(row=row_index, column=1, value=label)
        worksheet.cell(row=row_index, column=2, value=value)
        row_index += 1

    top_survey_start = row_index + 2
    worksheet.cell(row=top_survey_start, column=1, value="Top Survey")
    worksheet.cell(row=top_survey_start, column=1).fill = SUMMARY_HEADER_FILL
    worksheet.cell(row=top_survey_start, column=1).font = SUMMARY_HEADER_FONT
    worksheet.cell(row=top_survey_start, column=2, value="Total Histori")
    worksheet.cell(row=top_survey_start, column=2).fill = SUMMARY_HEADER_FILL
    worksheet.cell(row=top_survey_start, column=2).font = SUMMARY_HEADER_FONT

    for offset, survey_row in enumerate(rekap_survey_rows[:10], start=1):
        worksheet.cell(row=top_survey_start + offset, column=1, value=survey_row["nama_survei"])
        worksheet.cell(row=top_survey_start + offset, column=2, value=int(survey_row["total_histori"]))

    top_mitra_start = top_survey_start + max(len(rekap_survey_rows[:10]), 1) + 3
    headers = ["Top Mitra Berdasarkan Jumlah Histori", "Total Histori", "Total Survei Unik"]
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=top_mitra_start, column=column_index, value=header)
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = SUMMARY_HEADER_FONT

    for offset, mitra_row in enumerate(rekap_mitra_rows[:10], start=1):
        worksheet.cell(row=top_mitra_start + offset, column=1, value=mitra_row["nama_lengkap"])
        worksheet.cell(row=top_mitra_start + offset, column=2, value=int(mitra_row["total_histori"]))
        worksheet.cell(row=top_mitra_start + offset, column=3, value=int(mitra_row["total_survei_unik"]))

    autosize_worksheet_columns(worksheet)


def export_excel(bundle: ReportBundle, filename: str = "mitra_history_report.xlsx") -> Path:
    workbook = Workbook()
    write_summary_sheet(
        workbook,
        source_label=bundle.source_label,
        mitra_rows=bundle.mitra_rows,
        history_rows=bundle.history_rows,
        rekap_mitra_rows=bundle.rekap_mitra_rows,
        rekap_survey_rows=bundle.rekap_survey_rows,
    )
    write_table_sheet(workbook, "Mitra", bundle.mitra_rows)
    write_table_sheet(workbook, "History", bundle.history_rows)
    write_table_sheet(workbook, "Rekap Mitra", bundle.rekap_mitra_rows)
    write_table_sheet(workbook, "Rekap Survey", bundle.rekap_survey_rows)

    output_path = bundle.output_dir / filename
    bundle.output_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
